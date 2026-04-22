"""
Ontario Powerlifting Association provincial qualifying-total scraper.

OPA publishes its provincial qualifying totals as an Excel file on
Dropbox, linked from ``/qualifying-standards`` on ontariopowerlifting.org.
The Dropbox URL rotates whenever OPA re-uploads a revision, so the
scraper rediscovers the link from the HTML page at each run.

Sheet layout (Classic tab, as of 2026-04):
  - Row 1: "Men's Classic Provincial Qualifying Standards"
  - Row 2: ``Weight Class | 53kg | cpu | 59kg | (blank) | 66kg | ...``
  - Rows 3-9: one per age division
                (``Open, Sub-Junior, Junior, Master I..IV``)
  - Blank separator row
  - Repeated block for Women

The scraper filters the workbook to the Classic sheet, ignoring
Equipped and Bench which are out of scope for this project.
Missing-cell values show up as either ``"-"`` or ``None`` and both
translate to "no QT at this combination".
"""
from __future__ import annotations

import html
import logging
import re
from pathlib import Path
from typing import Iterable

import openpyxl
import requests

log = logging.getLogger(__name__)

OPA_LANDING_URL = "https://www.ontariopowerlifting.org/qualifying-standards"
_USER_AGENT = "cpu-analytics-scraper/1.0 (+contact matthias.bernhard7@gmail.com)"
_HTTP_TIMEOUT = 30.0

# Regex for the Dropbox-hosted Qualifying-Standards.xlsx link.
_DROPBOX_RE = re.compile(
    r'https?://www\.dropbox\.com/scl/fi/[^"\s]+/Qualifying-Standards\.xlsx[^"\s]*',
    re.IGNORECASE,
)

# Division label translation: Excel uses "Master I (40-49)", we
# canonicalise to "Master 1" etc.
_DIVISION_RE = re.compile(
    r"^(Open|Sub-?Junior|Junior|Master\s+(I{1,3}V?|IV))\b",
    re.IGNORECASE,
)
_MASTER_ROMAN = {"I": "Master 1", "II": "Master 2", "III": "Master 3", "IV": "Master 4"}


def _looks_like_decoded_html(text: str) -> bool:
    """Sanity check that we got a decoded HTML document, not a blob
    of compressed bytes that ``requests`` failed to decompress.

    OPA is hosted on a Wix CDN that serves ``Content-Encoding: br``
    regardless of Accept-Encoding. Without the ``Brotli`` Python
    package installed, ``requests.text`` silently returns garbage and
    the downstream regex misses the xlsx URL. This check turns that
    silent failure into a loud RuntimeError pointing at the root cause.
    """
    if not text:
        return False
    # HTML must start with a recognisable prefix. Allow a small amount
    # of leading whitespace / BOM. A real OPA page begins with
    # ``<!DOCTYPE html>``; failure mode is binary bytes that rarely
    # start with ``<``.
    head = text.lstrip()[:200].lower()
    return "<html" in head or "<!doctype" in head


def discover_xlsx_url(landing_url: str = OPA_LANDING_URL) -> str:
    """Fetch OPA's qualifying-standards page and return the current
    Dropbox URL to the Classic QT Excel workbook.

    Raises RuntimeError if the landing page contains no Dropbox link
    (OPA broke their site, or the link moved elsewhere), or if the
    response body is not decoded HTML (likely a missing decompressor).
    """
    r = requests.get(
        landing_url,
        headers={"User-Agent": _USER_AGENT},
        timeout=_HTTP_TIMEOUT,
    )
    r.raise_for_status()
    text = r.text
    if not _looks_like_decoded_html(text):
        encoding = r.headers.get("Content-Encoding", "")
        raise RuntimeError(
            f"OPA landing page body does not look like decoded HTML "
            f"(Content-Encoding={encoding!r}, len={len(text)}). The "
            f"Wix CDN serves Brotli and requests needs the Brotli "
            f"package installed to decode it. Install "
            f"``pip install Brotli`` or ensure backend/requirements.txt "
            f"pins it."
        )
    m = _DROPBOX_RE.search(text)
    if m is None:
        raise RuntimeError(
            f"no Dropbox .xlsx link found on {landing_url}; OPA likely "
            f"changed their page structure"
        )
    # Force direct download by setting dl=1 (Dropbox's viewer mode is
    # dl=0, which serves an HTML preview instead of the file). The
    # landing page HTML-escapes ``&`` in hrefs, so unescape first.
    url = html.unescape(m.group(0))
    url = re.sub(r"dl=0", "dl=1", url)
    if "dl=" not in url:
        url = url + ("&dl=1" if "?" in url else "?dl=1")
    return url


def extract_xlsx_url_from_html(text: str) -> str | None:
    """Pure-function variant of the URL discovery step. Used by tests
    to exercise the regex against a committed HTML fixture without
    hitting the network. Returns the post-escaping direct-download
    URL, or None if no match."""
    if not _looks_like_decoded_html(text):
        return None
    m = _DROPBOX_RE.search(text)
    if m is None:
        return None
    url = html.unescape(m.group(0))
    url = re.sub(r"dl=0", "dl=1", url)
    if "dl=" not in url:
        url = url + ("&dl=1" if "?" in url else "?dl=1")
    return url


def download_xlsx(url: str, target_dir: Path) -> Path:
    """Download the OPA xlsx to ``target_dir`` and return the local path."""
    target_dir.mkdir(parents=True, exist_ok=True)
    dst = target_dir / "opa_qualifying_standards.xlsx"
    r = requests.get(
        url,
        headers={"User-Agent": _USER_AGENT},
        timeout=_HTTP_TIMEOUT,
        stream=True,
    )
    r.raise_for_status()
    with dst.open("wb") as f:
        for chunk in r.iter_content(chunk_size=64 * 1024):
            if chunk:
                f.write(chunk)
    log.info("downloaded opa xlsx (%d bytes)", dst.stat().st_size)
    return dst


def _canonicalise_division(raw: str) -> str | None:
    """Map an OPA division label to our canonical vocabulary."""
    if not raw:
        return None
    raw = raw.strip()
    m = _DIVISION_RE.match(raw)
    if m is None:
        return None
    matched = m.group(1)
    if matched.lower().startswith("master"):
        # "Master I (40-49)" -> "Master 1"
        roman = re.sub(r"^[Mm]aster\s+", "", matched).upper()
        return _MASTER_ROMAN.get(roman)
    # "Sub-Junior" may lose the hyphen in the regex match
    if matched.lower().startswith("sub"):
        return "Sub-Junior"
    return matched.title()


def _weight_class_from_header(cell) -> str | None:
    """Turn a header cell like ``'83kg'`` or ``'120+kg'`` into ``'83'``
    or ``'120+'``. Returns None for cells that aren't a weight class."""
    if cell is None:
        return None
    s = str(cell).strip().lower()
    if not s.endswith("kg"):
        return None
    wc = s[:-2].strip()
    if not re.match(r"^\d+\+?$", wc):
        return None
    return wc


def _parse_qt(value) -> float | None:
    """Parse a QT cell. Returns None for blanks, dashes, or non-numeric."""
    if value is None:
        return None
    s = str(value).strip().replace("\xa0", "")  # nbsp artefact in the xlsx
    if s in ("", "-", "—", "–"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_xlsx(xlsx_path: Path) -> list[dict]:
    """
    Parse the OPA Classic sheet into QT rows.

    Scope: Classic + SBD only. Equipped and Bench sheets are skipped.
    Returns a list of row dicts matching ``data.scrapers.base.CSV_FIELDS``
    minus ``source_pdf`` / ``fetched_at`` (orchestrator adds those).
    """
    rows: list[dict] = []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Classic" not in wb.sheetnames:
        raise RuntimeError(
            f"OPA xlsx missing expected 'Classic' sheet; got {wb.sheetnames}"
        )
    ws = wb["Classic"]

    # Walk the sheet row-by-row. State machine:
    #   - Encountered a title row containing "Provincial Qualifying
    #     Standards" -> identify sex from text ("Men's" / "Women's").
    #   - Encountered a "Weight Class" header row -> record column ->
    #     weight-class mapping for subsequent data rows.
    #   - Encountered a data row (col 0 is a division label) -> emit.
    current_sex: str | None = None
    column_wc: dict[int, str] = {}

    for row in ws.iter_rows(values_only=True):
        if not row or all(c is None for c in row):
            continue
        first = row[0]
        first_s = str(first).strip() if first is not None else ""

        if "Provincial Qualifying Standards" in first_s:
            low = first_s.lower()
            if "women" in low:
                current_sex = "F"
            elif "men" in low:
                current_sex = "M"
            else:
                current_sex = None
            column_wc = {}
            continue

        if first_s.lower().startswith("weight class"):
            column_wc = {}
            for i, cell in enumerate(row[1:], start=1):
                wc = _weight_class_from_header(cell)
                if wc is not None:
                    column_wc[i] = wc
            continue

        division = _canonicalise_division(first_s)
        if division is None or current_sex is None or not column_wc:
            continue

        for i, wc in column_wc.items():
            val = row[i] if i < len(row) else None
            qt = _parse_qt(val)
            if qt is None:
                continue
            rows.append({
                "sex": current_sex,
                "level": "Provincials",
                "region": None,
                "division": division,
                "equipment": "Classic",
                "event": "SBD",
                "weight_class": wc,
                "qt": qt,
                "effective_year": 2026,
                "province": "Ontario",
            })

    log.info("parsed %d rows from OPA xlsx", len(rows))
    return rows


def parse_xlsxes(paths: Iterable[Path]) -> list[dict]:
    out: list[dict] = []
    for p in paths:
        out.extend(parse_xlsx(p))
    return out
